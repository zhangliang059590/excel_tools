# coding=utf-8
import copy
import random
import sys
import openpyxl as xl


def selector(fname):
    print("processing: %s" % fname, file=sys.stderr)
    wb = xl.load_workbook(fname)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        old_sheetname = "原-" + sheetname
        old_sheet = wb.copy_worksheet(sheet)
        old_sheet.title = old_sheetname
        show_rows = 0
        n_cols = sheet.max_column
        n_rows = sheet.max_row

        rand_seeds = []
        flag_index = -1
        for i in range(1, n_cols + 1):
            if sheet.cell(1, i).value.strip() == '是否生产':
                flag_index = i
        print("idx: %d" % flag_index, file=sys.stderr)
        if flag_index == -1:
            print("col error!", file=sys.stderr)
            exit(-1)
        for i in range(2, n_rows + 1):
            if sheet.row_dimensions[i].hidden == 1:
                continue
            flag_value = sheet.cell(i, flag_index).value.strip()
            if flag_value == '是':
                show_rows += 1
                rand_seeds.append(random.random())
        selected_num = int(show_rows * 0.3)
        seed_cp = copy.deepcopy(rand_seeds)
        seed_cp.sort()
        threshold = seed_cp[selected_num]
        print("row_num: %d; process_num: %d; \
                           selected_num: %d; threshold: % f"
              % (n_rows, show_rows, selected_num, threshold))
        j = 0
        sheet.cell(1, n_cols + 1, u'是否抽审')
        count = 0
        for i in range(2, n_rows + 1):
            if sheet.row_dimensions[i].hidden == 1:
                continue
            flag_value = sheet.cell(i, flag_index).value.strip()
            if flag_value != '是':
                continue
            if rand_seeds[j] < threshold:
                sheet.cell(i, n_cols + 1, u'是')
                count += 1
            else:
                sheet.cell(i, n_cols + 1, u'否')
            j += 1
        #for i in range(2, n_rows + 1):
        #    print("\rdeleting %d" % i, file=sys.stderr)
        #    if sheet.row_dimensions[i].hidden == 1:
        #        sheet.delete_rows(i)
        #for i in range(1, n_cols + 1):
        #    if sheet.col_dimensions[i].hidden == 1:
        #        sheet.delete_cols(i)
        print(sheetname, file=sys.stderr)
    out_name = fname.rsplit('.', 1)
    #wb.save(out_name[0] + "test.xlsx")
    wb.save(out_name[0] + (u'.抽审%d.output.' % count)+ out_name[1])


if __name__ == "__main__":
    selector(sys.argv[1])
